import datetime
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from database.db_manager import db
from services.theme_and_log_service import logger


class ReportService:
    @staticmethod
    def calculate_analytics():
        students = db.get_students()
        logs = db.get_all_logs()

        analytics = {}
        for student in students:
            analytics[student['id']] = {
                "student_number": student['student_number'],
                "name": f"{student['first_name']} {student['last_name']}",
                "homework_done": 0,
                "homework_missing": 0,
                "participation_net": 0,
                "behavior_net": 0
            }

        for log in logs:
            s_id = log['student_id']
            if s_id not in analytics:
                continue

            modifier = 1 if log['log_type'] in ["+", "Quick Score"] else -1
            category = log['category_tag']

            if category in ["participation", "Derse Katılım"]:
                analytics[s_id]["participation_net"] += modifier
            elif category in ["behavior", "Davranış"]:
                analytics[s_id]["behavior_net"] += modifier

        try:
            with db.get_connection() as conn:
                cursor = conn.cursor()
                cursor.execute("SELECT student_id, status FROM homework_checks")
                hw_checks = cursor.fetchall()

                for check in hw_checks:
                    s_id = check['student_id']
                    status = check['status']

                    if s_id in analytics:
                        if status == "Done":
                            analytics[s_id]["homework_done"] += 1
                        elif status == "Missing":
                            analytics[s_id]["homework_missing"] += 1
        except Exception as e:
            logger.error(f"Error fetching homework checks for report: {e}")

        return list(analytics.values())

    @staticmethod
    def export_to_styled_excel(class_name, file_path):
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = f"{class_name} Raporu"
            ws.views.sheetView[0].showGridLines = True

            font_title = Font(name="Segoe UI", size=16, bold=True, color="172B4D")
            font_sub = Font(name="Segoe UI", size=10, italic=True, color="5E6C84")
            font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
            font_body = Font(name="Segoe UI", size=10, color="172B4D")
            font_bold = Font(name="Segoe UI", size=10, bold=True, color="172B4D")

            fill_header = PatternFill(start_color="172B4D", end_color="172B4D", fill_type="solid")
            fill_zebra = PatternFill(start_color="F4F5F7", end_color="F4F5F7", fill_type="solid")
            fill_green = PatternFill(start_color="E3FCEF", end_color="E3FCEF", fill_type="solid")
            fill_red = PatternFill(start_color="FFEBE6", end_color="FFEBE6", fill_type="solid")

            align_center = Alignment(horizontal="center", vertical="center")
            align_left = Alignment(horizontal="left", vertical="center")

            thin_border = Border(
                left=Side(style='thin', color='DFE1E6'),
                right=Side(style='thin', color='DFE1E6'),
                top=Side(style='thin', color='DFE1E6'),
                bottom=Side(style='thin', color='DFE1E6')
            )

            ws.merge_cells("A1:F1")
            ws["A1"] = f"ÖĞRENCİ PERFORMANS VE ÖDEV RAPORU — {class_name}"
            ws["A1"].font = font_title
            ws["A1"].alignment = align_left

            now_str = datetime.datetime.now().strftime("%d.%m.%Y - %H:%M")
            ws.merge_cells("A2:F2")
            ws["A2"] = f"Rapor Oluşturulma Tarihi: {now_str} | Sistem: Class Tool"
            ws["A2"].font = font_sub
            ws["A2"].alignment = align_left

            ws.append([])

            headers = ["Öğrenci No", "Adı Soyadı", "Yapılan Ödev (✅)", "Eksik Ödev (❌)", "Net Derse Katılım", "Net Davranış"]
            ws.append(headers)
            header_row = 4

            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=header_row, column=col_num)
                cell.font = font_header
                cell.fill = fill_header
                cell.alignment = align_center
                cell.border = thin_border

            data = ReportService.calculate_analytics()
            start_data_row = 5

            for row_idx, s in enumerate(data, start=start_data_row):
                row_cells = [
                    ws.cell(row=row_idx, column=1, value=s['student_number']),
                    ws.cell(row=row_idx, column=2, value=s['name']),
                    ws.cell(row=row_idx, column=3, value=s['homework_done']),
                    ws.cell(row=row_idx, column=4, value=s['homework_missing']),
                    ws.cell(row=row_idx, column=5, value=s['participation_net']),
                    ws.cell(row=row_idx, column=6, value=s['behavior_net'])
                ]

                is_even = (row_idx % 2 == 0)

                for col_idx, cell in enumerate(row_cells, 1):
                    cell.font = font_body
                    cell.border = thin_border

                    if is_even:
                        cell.fill = fill_zebra

                    if col_idx == 1:
                        cell.alignment = align_center
                        cell.font = font_bold
                    elif col_idx == 2:
                        cell.alignment = align_left
                    elif col_idx in [3, 4, 5, 6]:
                        cell.alignment = align_center

                    if col_idx == 3 and s['homework_done'] > 0:
                        cell.fill = fill_green
                        cell.font = Font(name="Segoe UI", size=10, bold=True, color="006644")
                    elif col_idx == 4 and s['homework_missing'] > 0:
                        cell.fill = fill_red
                        cell.font = Font(name="Segoe UI", size=10, bold=True, color="BF2600")

            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    if cell.row in [1, 2]:
                        continue
                    if cell.value is not None:
                        max_len = max(max_len, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = max(max_len + 5, 14)

            wb.save(file_path)
            logger.info(f"Excel report saved successfully to {file_path}")
            return True
        except Exception as e:
            logger.error(f"Excel Export Crash Error: {e}", exc_info=True)
            raise e

    @staticmethod
    def export_student_pdf_report(student_id, file_path):
        """Belirli bir öğrencinin istatistiklerini, katılım netlerini, profil notlarını ve rozetlerini PDF yapısına dönüştürür."""
        try:
            import os
            from reportlab.lib.pagesizes import A4
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
            from services.badge_service import BadgeService

            font_path = "C:\\Windows\\Fonts\\arial.ttf"
            if os.path.exists(font_path):
                pdfmetrics.registerFont(TTFont('ArialCustom', font_path))
                pdfmetrics.registerFont(TTFont('ArialCustom-Bold', "C:\\Windows\\Fonts\\arialbd.ttf"))
                base_font = 'ArialCustom'
                base_font_bold = 'ArialCustom-Bold'
            else:
                base_font = 'Helvetica'
                base_font_bold = 'Helvetica-Bold'

            with db.get_connection() as conn:
                cursor = conn.cursor()
                cursor.execute("""
                               SELECT s.student_number,
                                      s.first_name,
                                      s.last_name,
                                      s.gender,
                                      s.class_id,
                                      c.name as class_name
                               FROM students s
                                        JOIN classes c ON s.class_id = c.id
                               WHERE s.id = ?
                               """, (student_id,))
                student = cursor.fetchone()
                if not student:
                    raise Exception("Öğrenci bulunamadı.")

                # Otomatik rozet hesaplamasını PDF basılmadan önce güncelleyelim
                BadgeService.calculate_auto_badges_for_class(student['class_id'])
                badges = BadgeService.get_student_badges(student_id)

                cursor.execute("SELECT * FROM student_profiles WHERE student_id = ?", (student_id,))
                profile = cursor.fetchone()

                cursor.execute("SELECT status FROM homework_checks WHERE student_id = ?", (student_id,))
                hw_checks = cursor.fetchall()
                total_hw = len(hw_checks)
                done_hw = sum(1 for c in hw_checks if c['status'] == 'Done')
                missing_hw = sum(1 for c in hw_checks if c['status'] == 'Missing')

                cursor.execute("SELECT log_type, category_tag FROM logs WHERE student_id = ?", (student_id,))
                logs = cursor.fetchall()
                part_net = sum(1 for l in logs if
                               l['log_type'] in ['+', 'Quick Score', 'Doğru'] and l['category_tag'] in ['participation',
                                                                                                        'Derse Katılım'])
                part_net -= sum(1 for l in logs if
                                l['log_type'] in ['-', 'Yanlış'] and l['category_tag'] in ['participation',
                                                                                           'Derse Katılım'])

                beh_net = sum(1 for l in logs if
                              l['log_type'] in ['+', 'Quick Score'] and l['category_tag'] in ['behavior', 'Davranış'])
                beh_net -= sum(
                    1 for l in logs if l['log_type'] in '-' and l['category_tag'] in ['behavior', 'Davranış'])

            gender_str = str(student['gender']).lower()
            if gender_str in ['female', 'kız', 'k']:
                primary_color = colors.HexColor('#DB2777')
                header_bg = colors.HexColor('#FCE7F3')
                table_header_bg = colors.HexColor('#BE185D')
                accent_color = colors.HexColor('#F59E0B')
                box_bg = colors.HexColor('#FFFBEB')
                box_border = colors.HexColor('#FCD34D')
            else:
                primary_color = colors.HexColor('#1E40AF')
                header_bg = colors.HexColor('#E0E7FF')
                table_header_bg = colors.HexColor('#1E3A8A')
                accent_color = colors.HexColor('#3B82F6')
                box_bg = colors.HexColor('#EFF6FF')
                box_border = colors.HexColor('#93C5FD')

            doc = SimpleDocTemplate(file_path, pagesize=A4, rightMargin=40, leftMargin=40, topMargin=40,
                                    bottomMargin=40)
            story = []
            styles = getSampleStyleSheet()

            title_style = ParagraphStyle('DocTitle', parent=styles['Heading1'], fontName=base_font_bold, fontSize=18,
                                         textColor=primary_color, alignment=1, spaceAfter=10)
            sub_style = ParagraphStyle('SubTitle', parent=styles['Normal'], fontName=base_font, fontSize=10,
                                       textColor=colors.HexColor('#4B5563'), alignment=1, spaceAfter=20)
            section_style = ParagraphStyle('SectionHeading', parent=styles['Heading2'], fontName=base_font_bold,
                                           fontSize=13, textColor=primary_color, spaceBefore=12, spaceAfter=6)
            body_style = ParagraphStyle('BodyDark', parent=styles['Normal'], fontName=base_font, fontSize=10,
                                        textColor=colors.HexColor('#1F2937'), leading=14)

            story.append(Paragraph("<b>ÖĞRENCİ BİREYSEL GELİŞİM RAPORU</b>", title_style))
            story.append(Paragraph(
                f"Sınıf: {student['class_name']} | Rapor Tarihi: {datetime.datetime.now().strftime('%d.%m.%Y')}",
                sub_style))

            data_info = [
                [Paragraph("<b>Öğrenci No:</b>", body_style), Paragraph(str(student['student_number']), body_style)],
                [Paragraph("<b>Adı Soyadı:</b>", body_style),
                 Paragraph(f"{student['first_name']} {student['last_name']}", body_style)],
                [Paragraph("<b>Cinsiyet:</b>", body_style), Paragraph(str(student['gender']), body_style)]
            ]
            t_info = Table(data_info, colWidths=[120, 390])
            t_info.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, -1), header_bg),
                ('BOX', (0, 0), (-1, -1), 1, accent_color),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('PADDING', (0, 0), (-1, -1), 7),
            ]))
            story.append(t_info)
            story.append(Spacer(1, 10))

            # ROZETLER BÖLÜMÜ (PDF UYUMLU METİN ETİKETLERİ)
            if badges:
                story.append(Paragraph("<b>Kazanılan Başarı Rozetleri</b>", section_style))
                badge_text_list = [f"<b>[{b['title']}]</b>: {b['desc']}" for b in badges]
                data_b = [[Paragraph(b_txt, body_style)] for b_txt in badge_text_list]
                t_b = Table(data_b, colWidths=[510])
                t_b.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, -1), header_bg),
                    ('BOX', (0, 0), (-1, -1), 1, accent_color),
                    ('PADDING', (0, 0), (-1, -1), 6),
                ]))
                story.append(t_b)
                story.append(Spacer(1, 10))

            story.append(Paragraph("<b>Akademik ve Ders İçi Performans Özetleri</b>", section_style))
            data_stats = [
                [Paragraph(f"<font color='white'><b>Metrik / Kriter</b></font>", body_style),
                 Paragraph(f"<font color='white'><b>İstatistik Değeri</b></font>", body_style)],
                [Paragraph("Toplam Atanan / Yapılan Ödev", body_style),
                 Paragraph(f"Yapılan: {done_hw} / Toplam: {total_hw} (Eksik: {missing_hw})", body_style)],
                [Paragraph("Net Derse Katılım Puanı", body_style), Paragraph(str(part_net), body_style)],
                [Paragraph("Net Davranış Puanı", body_style), Paragraph(str(beh_net), body_style)],
            ]
            t_stats = Table(data_stats, colWidths=[250, 260])
            t_stats.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), table_header_bg),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#D1D5DB')),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('PADDING', (0, 0), (-1, -1), 6),
            ]))
            story.append(t_stats)
            story.append(Spacer(1, 10))

            if profile:
                story.append(Paragraph("<b>Kişilik Etiketleri ve Profil Puanları</b>", section_style))
                soc = profile['sociability_score'] or 3
                foc = profile['focus_score'] or 3
                part = profile['participation_score'] or 3
                tags = profile['personality_tags'] or "Belirtilmemiş"
                notes = profile['teacher_notes'] or "Öğretmen özel notu girilmemiş."

                data_profile = [
                    [Paragraph("<b>Sosyallik (1-5):</b>", body_style), Paragraph(str(soc), body_style)],
                    [Paragraph("<b>Odaklanma (1-5):</b>", body_style), Paragraph(str(foc), body_style)],
                    [Paragraph("<b>Katılım (1-5):</b>", body_style), Paragraph(str(part), body_style)],
                    [Paragraph("<b>Karakteristik Etiketler:</b>", body_style), Paragraph(tags, body_style)],
                ]
                t_prof = Table(data_profile, colWidths=[150, 360])
                t_prof.setStyle(TableStyle([
                    ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#D1D5DB')),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ('PADDING', (0, 0), (-1, -1), 6),
                ]))
                story.append(t_prof)
                story.append(Spacer(1, 10))

                story.append(Paragraph("<b>Öğretmen Gözlem Notları:</b>", section_style))
                data_notes = [[Paragraph(notes, body_style)]]
                t_notes = Table(data_notes, colWidths=[510])
                t_notes.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, -1), box_bg),
                    ('BOX', (0, 0), (-1, -1), 1, box_border),
                    ('PADDING', (0, 0), (-1, -1), 8),
                ]))
                story.append(t_notes)

            doc.build(story)
            return True
        except Exception as e:
            raise Exception(f"PDF Raporu Oluşturulamadı: {str(e)}")